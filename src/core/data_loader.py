"""
DataLoader - 测试数据加载器

支持加载 Excel (.xlsx) 和 CSV 格式的测试数据文件。
提供数据摘要生成功能，供 LLM Prompt 使用。
"""

import csv
import logging
from pathlib import Path
from typing import Dict, List, Any, Optional

logger = logging.getLogger(__name__)


class DataLoader:
    """测试数据加载器

    支持格式:
    - Excel (.xlsx): 多 Sheet 支持，每个 Sheet 作为独立数据集
    - CSV (.csv): 单文件单数据集
    """

    @staticmethod
    def load(file_path: str) -> Dict[str, List[Dict[str, Any]]]:
        """加载测试数据文件

        Args:
            file_path: 文件路径 (支持 .xlsx 或 .csv)

        Returns:
            Dict[str, List[Dict]]: {数据集名称: [行数据字典, ...]}
            - Excel: 数据集名称为 Sheet 名
            - CSV: 数据集名称为文件名 (不含扩展名)

        Raises:
            ValueError: 不支持的文件格式
            FileNotFoundError: 文件不存在
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"测试数据文件不存在: {file_path}")

        ext = path.suffix.lower()
        if ext == '.xlsx':
            return DataLoader.load_excel(file_path)
        elif ext == '.csv':
            return {path.stem: DataLoader.load_csv(file_path)}
        else:
            raise ValueError(f"不支持的测试数据格式: {ext}，仅支持 .xlsx 和 .csv")

    @staticmethod
    def load_excel(file_path: str) -> Dict[str, List[Dict[str, Any]]]:
        """加载 Excel 文件

        Args:
            file_path: Excel 文件路径

        Returns:
            Dict[str, List[Dict]]: {Sheet名: [行数据字典, ...]}
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("需要安装 openpyxl: pip install openpyxl")

        logger.info(f"加载 Excel 文件: {file_path}")
        wb = openpyxl.load_workbook(file_path, data_only=True)
        result = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            if not rows:
                logger.debug(f"跳过空 Sheet: {sheet_name}")
                continue

            # 第一行作为表头
            headers = [str(cell) if cell is not None else f"column_{i}"
                       for i, cell in enumerate(rows[0])]

            # 过滤掉全空的列
            valid_cols = [i for i, h in enumerate(headers)
                          if not h.startswith("column_")]

            data_rows = []
            for row in rows[1:]:
                # 跳过空行
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    continue

                row_dict = {}
                for i in valid_cols:
                    if i < len(row):
                        value = row[i]
                        # 处理特殊类型
                        if value is not None:
                            row_dict[headers[i]] = DataLoader._normalize_value(value)
                        else:
                            row_dict[headers[i]] = None

                if row_dict:
                    data_rows.append(row_dict)

            if data_rows:
                result[sheet_name] = data_rows
                logger.info(f"  Sheet '{sheet_name}': {len(data_rows)} 行, 列: {headers[:5]}...")

        wb.close()
        return result

    @staticmethod
    def load_csv(file_path: str, encoding: str = 'utf-8') -> List[Dict[str, Any]]:
        """加载 CSV 文件

        Args:
            file_path: CSV 文件路径
            encoding: 文件编码，默认 UTF-8

        Returns:
            List[Dict]: [行数据字典, ...]
        """
        logger.info(f"加载 CSV 文件: {file_path}")

        # 尝试使用 chardet 自动检测编码
        detected_encoding = None
        try:
            import chardet
            with open(file_path, 'rb') as f:
                raw = f.read(10000)
                detected = chardet.detect(raw)
                if detected and detected.get('confidence', 0) > 0.7:
                    detected_encoding = detected['encoding']
                    logger.info(f"自动检测编码: {detected_encoding} (置信度: {detected['confidence']:.0%})")
        except ImportError:
            pass  # chardet 未安装，使用 fallback 列表

        # 构建编码尝试列表（检测到的编码优先）
        encodings = [encoding, 'utf-8-sig', 'gbk', 'gb2312', 'gb18030', 'big5', 'cp936', 'latin-1']
        if detected_encoding:
            encodings.insert(0, detected_encoding)
        # 去重保持顺序
        seen = set()
        encodings = [x for x in encodings if not (x in seen or seen.add(x))]

        for enc in encodings:
            try:
                with open(file_path, 'r', encoding=enc, newline='') as f:
                    # 自动检测分隔符
                    sample = f.read(4096)
                    f.seek(0)

                    # 尝试使用 Sniffer 检测分隔符，失败时使用默认逗号
                    try:
                        dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
                        reader = csv.DictReader(f, dialect=dialect)
                    except csv.Error:
                        # Sniffer 失败（可能是数据中有未转义的逗号），使用默认逗号分隔
                        logger.debug(f"Sniffer 无法确定分隔符，使用默认逗号分隔")
                        f.seek(0)
                        reader = csv.DictReader(f, delimiter=',')

                    rows = []
                    for row in reader:
                        # 规范化值
                        normalized_row = {
                            k: DataLoader._normalize_value(v)
                            for k, v in row.items()
                            if k is not None
                        }
                        if any(v is not None and v != "" for v in normalized_row.values()):
                            rows.append(normalized_row)

                    logger.info(f"  CSV: {len(rows)} 行")
                    return rows

            except UnicodeDecodeError:
                continue

        raise ValueError(f"无法解析 CSV 文件，尝试的编码: {encodings}")

    @staticmethod
    def _normalize_value(value: Any) -> Any:
        """规范化单元格值

        Args:
            value: 原始值

        Returns:
            规范化后的值
        """
        if value is None:
            return None

        # 字符串处理
        if isinstance(value, str):
            value = value.strip()
            if value == "":
                return None

            # 尝试转换为数字
            try:
                if '.' in value:
                    return float(value)
                return int(value)
            except ValueError:
                pass

            # 布尔值
            if value.lower() in ('true', 'yes', '是', '1'):
                return True
            if value.lower() in ('false', 'no', '否', '0'):
                return False

            return value

        # 数字保持原样
        if isinstance(value, (int, float)):
            return value

        # 其他类型转字符串
        return str(value)

    @staticmethod
    def summarize_for_prompt(
        data: Dict[str, List[Dict[str, Any]]],
        max_rows_preview: int = 3
    ) -> str:
        """生成数据摘要供 LLM Prompt 使用

        Args:
            data: 加载的数据 {数据集名: [行数据, ...]}
            max_rows_preview: 每个数据集预览的最大行数

        Returns:
            格式化的摘要文本
        """
        if not data:
            return "无测试数据"

        lines = []
        total_rows = 0

        for dataset_name, rows in data.items():
            if not rows:
                continue

            columns = list(rows[0].keys())
            total_rows += len(rows)

            lines.append(f"\n### 数据集: {dataset_name}")
            lines.append(f"- 行数: {len(rows)}")
            lines.append(f"- 列: {columns}")

            # 预览前几行
            if rows and max_rows_preview > 0:
                lines.append("- 数据预览:")
                for i, row in enumerate(rows[:max_rows_preview]):
                    preview_items = [f"{k}={v}" for k, v in list(row.items())[:5]]
                    lines.append(f"  [{i+1}] {', '.join(preview_items)}")
                if len(rows) > max_rows_preview:
                    lines.append(f"  ... 还有 {len(rows) - max_rows_preview} 行")

        summary = f"共 {len(data)} 个数据集, {total_rows} 行数据\n"
        summary += "\n".join(lines)
        return summary

    @staticmethod
    def load_multiple(file_paths: List[str]) -> Dict[str, List[Dict[str, Any]]]:
        """加载多个测试数据文件

        Args:
            file_paths: 文件路径列表

        Returns:
            合并后的数据 {数据集名: [行数据, ...]}
        """
        all_data = {}

        for file_path in file_paths:
            try:
                data = DataLoader.load(file_path)
                # 合并数据，如果有同名数据集则添加文件名前缀
                for name, rows in data.items():
                    if name in all_data:
                        # 添加文件名前缀避免冲突
                        file_stem = Path(file_path).stem
                        name = f"{file_stem}_{name}"
                    all_data[name] = rows
            except Exception as e:
                logger.warning(f"加载文件失败 {file_path}: {e}")

        return all_data

    @staticmethod
    def infer_api_mapping(
        data: Dict[str, List[Dict[str, Any]]],
        swagger_endpoints: List[Dict[str, Any]]
    ) -> Dict[str, Dict[str, Any]]:
        """推断测试数据与 API 接口的映射关系

        基于列名与接口参数名的相似度进行匹配。

        Args:
            data: 加载的测试数据
            swagger_endpoints: Swagger 接口列表

        Returns:
            映射关系 {接口路径: {dataset, columns, match_score}}
        """
        mapping = {}

        for dataset_name, rows in data.items():
            if not rows:
                continue

            columns = set(rows[0].keys())
            columns_lower = {c.lower() for c in columns}

            best_match = None
            best_score = 0

            for endpoint in swagger_endpoints:
                path = endpoint.get('path', '')
                method = endpoint.get('method', 'get')
                params = set()

                # 收集所有参数名
                for param in endpoint.get('parameters', []):
                    params.add(param.get('name', '').lower())

                # 收集 requestBody 的属性名
                if 'requestBody' in endpoint:
                    schema = endpoint['requestBody'].get('schema', {})
                    if 'properties' in schema:
                        params.update(k.lower() for k in schema['properties'].keys())

                if not params:
                    continue

                # 计算匹配分数 (交集/并集)
                intersection = columns_lower & params
                union = columns_lower | params
                score = len(intersection) / len(union) if union else 0

                if score > best_score and score > 0.3:  # 阈值 30%
                    best_score = score
                    best_match = {
                        'api': f"{method.upper()} {path}",
                        'matched_columns': list(intersection),
                        'match_score': round(score, 2)
                    }

            if best_match:
                api_key = best_match['api']
                if api_key not in mapping or mapping[api_key]['match_score'] < best_score:
                    mapping[api_key] = {
                        'dataset': dataset_name,
                        'columns': list(columns),
                        'row_count': len(rows),
                        **best_match
                    }

        return mapping
